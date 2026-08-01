"""
results_xlsx.py
===============
Four-sheet Excel results writer for DeBAIT fine-tuning pipeline.

"""

from __future__ import annotations
from typing import Dict, List, Optional, Set, Tuple
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. STYLING CONSTANTS
# =============================================================================

# Base Styles
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="1F4E79", size=11)
_KEY_ROW_FILL = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
_KEY_ROW_FONT = Font(size=7, color="AAAAAA")

# Delta Styles
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(color="006100")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT = Font(color="9C0006")
_NEUTRAL_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
_NEUTRAL_FONT = Font(color="9C6500")

# Layout Styles
_ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
_SECTION_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"), right=Side(style="medium", color="1F4E79"))
_ALIGN_CENTER = Alignment(wrap_text=False, vertical="center", horizontal="center")
_ALIGN_LEFT = Alignment(wrap_text=False, vertical="center", horizontal="left")
_ALIGN_WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
_MAX_COL_WIDTH = 42

# Vertical Sheet Specific Styles
_VERT_SECTION_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_VERT_SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)
_VERT_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
_VERT_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_VERT_RUN_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_VERT_RUN_HEADER_FONT = Font(bold=True, color="1F4E79", size=10)
_VERT_METRIC_FONT = Font(bold=False, size=10)
_VERT_METRIC_FONT_BOLD = Font(bold=True, size=10)
_VERT_DIRECTION_FONT = Font(italic=True, color="666666", size=9)

# Row/Col Indices
_ROW_SECTION, _ROW_DISPLAY, _ROW_KEYS, _ROW_DATA = 1, 2, 3, 4
_VERT_NUM_COLS = 6

# Loss function triggers for conditional config display
_ENTROPY_LOSSES = {"ce_dpo_kl_entropy", "ce_dpo_kl_all"}
_CONDITIONED_LOSSES = {"ce_dpo_kl_conditioned", "ce_dpo_kl_all"}
_CF_LOSSES = {"ce_dpo_kl_counterfactual", "ce_dpo_kl_all"}


# =============================================================================
# 2. METRIC CLASSIFICATION & ORDERING
# =============================================================================

_IDEAL_AT_50: Set[str] = {"SS", "pct_stereo"}
_LOWER_IS_BETTER: Set[str] = {
    "pct_negative", "regard_disparity", "fnr_toxic", "accuracy_disparity",
    "fnr_disparity", "parse_fail_rate", "bias_rate", "bias_disparity", "hedge_rate"
}
_HIGHER_IS_BETTER: Set[str] = {
    "LMS", "ICAT", "accuracy", "pct_anti", "pct_positive", "pct_neutral",
    "regard_mean", "macro_f1", "toxic_f1", "toxic_precision", "toxic_recall", "benign_f1"
}
_METADATA: Set[str] = {
    "total_prompts", "n_domains", "n_groups", "total", "total_lms", "total_ss",
    "n_parsed", "eval_mode", "n_biased", "n_categories", "n_dimensions", "n_groups_demo"
}

# Preferred column order within each dataset section
_ORDER_MAP: Dict[str, List[str]] = {
    "StereoSet": ["LMS", "ICAT", "SS"],
    "CrowS-Pairs": ["pct_stereo", "pct_anti"],
    "BOLD": ["regard_mean", "regard_disparity", "pct_positive", "pct_neutral", "pct_negative"],
    "ToxiGen": ["accuracy", "macro_f1", "toxic_f1", "toxic_precision", "toxic_recall", "benign_f1", "fnr_toxic", "accuracy_disparity", "fnr_disparity", "parse_fail_rate"],
    "BBQ": ["accuracy", "acc_disambig", "acc_ambig", "s_dis", "s_amb", "parse_fail_rate"],
    "ImplicitBBQ": ["accuracy", "hedge_rate", "parse_fail_rate"],
    "FairMTBench": ["bias_rate", "bias_disparity", "parse_fail_rate"],
}


# =============================================================================
# 3. READABLE NAMES & HELPERS
# =============================================================================

_READABLE_METRIC_NAMES: Dict[str, str] = {
    "accuracy": "Accuracy", "macro_f1": "Macro F1", "toxic_f1": "Toxic F1",
    "toxic_precision": "Toxic Precision", "toxic_recall": "Toxic Recall", "benign_f1": "Benign F1",
    "fnr_toxic": "False Negative Rate (toxic)", "accuracy_disparity": "Accuracy Disparity",
    "fnr_disparity": "FNR Disparity", "parse_fail_rate": "Parse Failure Rate",
    "regard_mean": "Mean Regard", "regard_disparity": "Regard Disparity",
    "pct_positive": "Positive %", "pct_neutral": "Neutral %", "pct_negative": "Negative %",
    "bias_rate": "Bias Rate (FairMT)", "bias_disparity": "Cross-Category Bias Disparity",
    "hedge_rate": "Hedge Rate (ImplicitBBQ)",
    "ce": "Cross-Entropy Loss", "total": "Total Loss", "dpo": "DPO Loss", "kl": "KL Regularisation",
    "hinge": "Hinge Loss", "entropy": "Entropy-Conditioned Loss (P1)",
    "dpo_conditioned": "Condition-Weighted DPO Loss (P2)", "cf": "Counterfactual Consistency Loss (P3)",
    "train_instructions": "Training Instructions", "eval_prompts": "Evaluation Prompts",
}

def _make_readable_metric_name(raw_metric: str) -> str:
    if raw_metric in _READABLE_METRIC_NAMES:
        return _READABLE_METRIC_NAMES[raw_metric]
    
    # Handle prefixed variants
    prefix_map = {
        "ss_": "SS", "stereo_": "Stereo", "regard_mean_": "Regard Mean",
        "disparity_": "Disparity", "acc_": "Accuracy", "fnr_": "FNR",
        "bias_rate_": "Bias Rate", "eval_BBQ_": "BBQ", "eval_BOLD_": "BOLD",
        "eval_ToxiGen_": "ToxiGen", "eval_FairMT_": "FairMT", "eval_ImplicitBBQ_": "ImplicitBBQ"
    }
    for prefix, label in prefix_map.items():
        if raw_metric.startswith(prefix):
            suffix = raw_metric[len(prefix):].replace("_", " ").title()
            return f"{label} — {suffix}" if suffix else label
            
    return raw_metric.replace("_", " ").title()


def _classify_metric(key: str) -> str:
    raw = _strip_prefix(key)
    if raw in _METADATA: return "unknown"
    if raw in _IDEAL_AT_50 or raw.startswith(("ss_", "stereo_")): return "ideal_at_50"
    if raw in _LOWER_IS_BETTER or raw.startswith(("disparity_", "fnr_")): return "良ower_is_better" # Typo fix in original logic
    if raw in _HIGHER_IS_BETTER or raw.startswith(("acc_", "regard_mean_")): return "higher_is_better"
    if raw.startswith("bias_rate_") and raw != "bias_rate": return "lower_is_better"
    return "unknown"

# Fix for the typo introduced in thought process above, ensuring correct string
def _classify_metric(key: str) -> str:
    raw = _strip_prefix(key)
    if raw in _METADATA: return "unknown"
    if raw in _IDEAL_AT_50 or raw.startswith(("ss_", "stereo_")): return "ideal_at_50"
    if raw in _LOWER_IS_BETTER or raw.startswith(("disparity_", "fnr_")): return "lower_is_better"
    if raw in _HIGHER_IS_BETTER or raw.startswith(("acc_", "regard_mean_")): return "higher_is_better"
    if raw.startswith("bias_rate_") and raw != "bias_rate": return "lower_is_better"
    return "unknown"


def _strip_prefix(key: str) -> str:
    for pfx in ("delta_", "baseline_", "postft_"):
        if key.startswith(pfx):
            key = key[len(pfx):]
            break
    parts = key.split("/")
    return "/".join(parts[2:]) if len(parts) >= 3 else (parts[1] if len(parts) == 2 else key)


def _extract_dataset_condition(key: str) -> Tuple[str, Optional[str], str]:
    raw = key
    for pfx in ("delta_", "baseline_", "postft_"):
        if raw.startswith(pfx):
            raw = raw[len(pfx):]
            break
    parts = raw.split("/")
    if len(parts) >= 3:
        return parts[0], parts[1], "/".join(parts[2:])
    elif len(parts) == 2:
        return parts[0], None, parts[1]
    return "Other", None, raw


def _is_improvement(delta_key: str, delta_value, baseline_value=None) -> Optional[bool]:
    try:
        d = float(delta_value)
    except (TypeError, ValueError):
        return None
    if abs(d) < 1e-6:
        return None

    cls = _classify_metric(delta_key)
    if cls == "ideal_at_50":
        try:
            base = float(baseline_value)
            dist_before, dist_after = abs(base - 50.0), abs((base + d) - 50.0)
            if dist_after < dist_before: return True
            if dist_after > dist_before: return False
        except (TypeError, ValueError):
            pass
        return None
    elif cls == "lower_is_better":
        return d < 0
    elif cls == "higher_is_better":
        return d > 0
    return None


def _direction_label(metric_key: str) -> str:
    cls = _classify_metric(metric_key)
    if cls == "ideal_at_50": return "→50 = ideal"
    if cls == "lower_is_better": return "↓ = better"
    if cls == "higher_is_better": return "↑ = better"
    return ""


def _format_prompt_list(pipe_separated: str) -> str:
    if not pipe_separated: return "(none)"
    items = [s.strip() for s in str(pipe_separated).split("|") if s.strip()]
    return ", ".join(items) if items else "(none)"


# =============================================================================
# 4. EXCEL FORMATTING HELPERS
# =============================================================================

def _autofit_columns(ws, max_width: int = _MAX_COL_WIDTH) -> None:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _freeze_panes(ws, row: int, col: int) -> None:
    ws.freeze_panes = f"{get_column_letter(col)}{row}"


def _apply_alt_row_shading(ws, data_start_row: int) -> None:
    merged_skip = {(r, c) for rng in ws.merged_cells.ranges for r in range(rng.min_row, rng.max_row + 1) for c in range(rng.min_col, rng.max_col + 1) if (r, c) != (rng.min_row, rng.min_col)}
    for row_idx in range(data_start_row, ws.max_row + 1):
        if (row_idx - data_start_row) % 2 == 1:
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) not in merged_skip:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.fill is None or cell.fill.fill_type is None:
                        cell.fill = _ALT_ROW_FILL


def _round_value(v, decimals: int = 4):
    return round(v, decimals) if isinstance(v, float) else v


def _make_display_header(col_name: str) -> str:
    for pfx in ("baseline_", "postft_", "delta_"):
        if col_name.startswith(pfx):
            raw = col_name[len(pfx):]
            parts = raw.split("/")
            metric_name = "/".join(parts[2:]) if len(parts) >= 3 else (parts[1] if len(parts) == 2 else raw)
            return f"{metric_name}\n({pfx.rstrip('_')})"
    return _READABLE_METRIC_NAMES.get(col_name, col_name)


# =============================================================================
# 5. SECTION DEFINITIONS (Sheet 1 & 3)
# =============================================================================

_SECTION_RUN_INFO = ["timestamp_utc", "hostname", "run_id", "wall_time_s"]
_SECTION_CONFIG_BASE = [
    "model_id", "train_dataset", "eval_datasets", "bbq_categories", "n_samples", "loss_fn",
    "bias_lambda", "hinge_lambda", "bias_margin", "dpo_beta", "dpo_ref", "kl_lambda", "kl_mode",
    "label_smooth", "focal_gamma", "epochs", "lr", "batch_size", "use_instruction", "use_cot",
    "per_epoch_eval", "stereo_from_metadata", "biasdpo_dir", "civil_comments_path",
    "civil_toxicity_threshold", "bold_path", "toxigen_path", "toxigen_threshold", "device",
]
_SECTION_CONFIG_PROMPTS = ["train_instructions", "eval_prompts"]
_SECTION_CONFIG_FAIRMT = ["fairmtbench_path", "fairmtbench_categories", "fairmtbench_judge_model", "fairmtbench_max_new_tokens"]
_SECTION_CONFIG_PROPOSALS = ["entropy_lambda", "cond_w_ambig_neg", "cond_w_ambig_nonneg", "cond_w_disambig_neg", "cond_w_disambig_nonneg", "cf_lambda", "cf_margin"]
_SECTION_CONFIG = _SECTION_CONFIG_BASE + _SECTION_CONFIG_PROMPTS + _SECTION_CONFIG_FAIRMT + _SECTION_CONFIG_PROPOSALS
_SECTION_TRAINING = ["n_train_samples", "n_biased_samples"]


# =============================================================================
# 6. DATA STRUCTURING HELPERS
# =============================================================================

def _build_triplet_columns(row: Dict) -> List[Dict]:
    delta_keys = sorted(k for k in row if k.startswith("delta_") and row[k] != "")
    triplets, covered = [], set()

    for dk in delta_keys:
        raw_full = dk[len("delta_"):]
        ds, condition, raw = _extract_dataset_condition(dk)
        triplets.append({
            "metric": raw_full, "raw": raw, "condition": condition, "dataset": ds,
            "section_key": f"{ds} ({condition})" if condition else ds,
            "baseline_key": f"baseline_{raw_full}", "postft_key": f"postft_{raw_full}", "delta_key": dk,
        })
        covered.update([f"baseline_{raw_full}", f"postft_{raw_full}"])

    for k in sorted(row.keys()):
        if k.startswith("baseline_") and k not in covered:
            raw_full = k[len("baseline_"):]
            ds, condition, raw = _extract_dataset_condition(k)
            triplets.append({
                "metric": raw_full, "raw": raw, "condition": condition, "dataset": ds,
                "section_key": f"{ds} ({condition})" if condition else ds,
                "baseline_key": k, "postft_key": f"postft_{raw_full}" if f"postft_{raw_full}" in row else None, "delta_key": None,
            })
            covered.add(k)
            if f"postft_{raw_full}" in row: covered.add(f"postft_{raw_full}")
    return triplets


def _build_ordered_sections(triplets: List[Dict], row: Dict) -> Tuple[List[Dict], List[str]]:
    section_groups: Dict[str, List[Dict]] = {}
    for t in triplets:
        section_groups.setdefault(t["section_key"], []).append(t)

    ds_order = ["StereoSet", "CrowS-Pairs", "BOLD", "ToxiGen", "BBQ", "ImplicitBBQ", "FairMTBench"]
    condition_order = ["plain", "reasoning"]
    ordered_section_keys, seen = [], set()

    for ds in ds_order:
        for cond in condition_order:
            sk = f"{ds} ({cond})"
            if sk in section_groups and sk not in seen:
                ordered_section_keys.append(sk)
                seen.add(sk)
        if ds in section_groups and ds not in seen:
            ordered_section_keys.append(ds)
            seen.add(ds)
            
    for sk in sorted(section_groups.keys()):
        if sk not in seen:
            ordered_section_keys.append(sk)
            seen.add(sk)

    ordered_sections = []
    for sk in ordered_section_keys:
        if sk not in section_groups: continue
        ds = section_groups[sk][0]["dataset"]
        ordered_sections.append({
            "name": sk,
            "triplets": sorted(section_groups[sk], key=lambda t: _metric_sort_key(t["raw"], ds)),
        })
    return ordered_sections, ordered_section_keys


def _metric_sort_key(raw_metric: str, dataset: str) -> tuple:
    order = _ORDER_MAP.get(dataset, [])
    if raw_metric in order:
        return (0, order.index(raw_metric), raw_metric)
    
    # Fallback prefix sorting
    prefix_checks = {
        "BOLD": ("regard_mean_", "disparity_"),
        "ToxiGen": ("acc_", "fnr_"),
        "FairMTBench": ("bias_rate_",),
    }
    for i, prefix in enumerate(prefix_checks.get(dataset, [])):
        if raw_metric.startswith(prefix):
            return (1, i, raw_metric)
    return (2, 0, raw_metric)


# =============================================================================
# 7. SHEET BUILDERS
# =============================================================================

def _read_key_row(ws) -> Dict[str, int]:
    return {str(ws.cell(row=_ROW_KEYS, column=ci).value): ci for ci in range(1, ws.max_column + 1) if ws.cell(row=_ROW_KEYS, column=ci).value}


def _extend_sheet_with_new_columns(ws, existing_key_map: Dict[str, int], new_columns: List[str], section_name: str = "New") -> Dict[str, int]:
    key_map = dict(existing_key_map)
    next_col = max(key_map.values(), default=0) + 1

    for col_name in new_columns:
        if col_name not in key_map:
            ci = next_col
            ws.cell(row=_ROW_SECTION, column=ci, value=section_name).font, ws.cell(row=_ROW_SECTION, column=ci).fill, ws.cell(row=_ROW_SECTION, column=ci).alignment = _SECTION_FONT, _SECTION_FILL, _ALIGN_CENTER
            ws.cell(row=_ROW_DISPLAY, column=ci, value=_make_display_header(col_name)).font, ws.cell(row=_ROW_DISPLAY, column=ci).fill, ws.cell(row=_ROW_DISPLAY, column=ci).alignment = _HEADER_FONT, _HEADER_FILL, _ALIGN_WRAP_CENTER
            ws.cell(row=_ROW_KEYS, column=ci, value=col_name).font, ws.cell(row=_ROW_KEYS, column=ci).fill, ws.cell(row=_ROW_KEYS, column=ci).alignment = _KEY_ROW_FONT, _KEY_ROW_FILL, _ALIGN_CENTER
            key_map[col_name] = ci
            next_col += 1
    return key_map


def _write_results_sheet(wb: Workbook, row: Dict) -> None:
    sheet_name, all_columns, sections = "Results", [], []
    triplets = _build_triplet_columns(row)
    baseline_lookup = {t["delta_key"]: float(row.get(t["baseline_key"], 0)) for t in triplets if t["delta_key"] and t["baseline_key"]}

    # Build sections
    for name, cols in [("Run Info", _SECTION_RUN_INFO), ("Config", _SECTION_CONFIG), ("Training", _SECTION_TRAINING)]:
        filtered = [c for c in cols if c in row]
        if filtered: sections.append({"name": name, "columns": filtered})
        
    ordered_metric_sections, _ = _build_ordered_sections(triplets, row)
    for sec in ordered_metric_sections:
        cols = [c for t in sec["triplets"] for c in (t["baseline_key"], t["postft_key"], t["delta_key"]) if c and c in row]
        if cols: sections.append({"name": sec["name"], "columns": cols})

    all_columns = [c for sec in sections for c in sec["columns"]]
    is_new = sheet_name not in wb.sheetnames

    if is_new:
        ws = wb.create_sheet(sheet_name, 0)
        col_map = {col: i for i, col in enumerate(all_columns, 1)}
        
        # Merge section headers
        for sec in sections:
            positions = [col_map[c] for c in sec["columns"] if c in col_map]
            if not positions: continue
            start, end = min(positions), max(positions)
            cell = ws.cell(row=_ROW_SECTION, column=start, value=sec["name"])
            cell.font, cell.fill, cell.alignment = _SECTION_FONT, _SECTION_FILL, _ALIGN_CENTER
            if start != end:
                ws.merge_cells(start_row=_ROW_SECTION, start_column=start, end_row=_ROW_SECTION, end_column=end)
                ws.cell(row=_ROW_SECTION, column=end).border = _SECTION_BORDER

        for col, ci in col_map.items():
            ws.cell(row=_ROW_DISPLAY, column=ci, value=_make_display_header(col)).font, ws.cell(row=_ROW_DISPLAY, column=ci).fill, ws.cell(row=_ROW_DISPLAY, column=ci).alignment, ws.cell(row=_ROW_DISPLAY, column=ci).border = _HEADER_FONT, _HEADER_FILL, _ALIGN_WRAP_CENTER, _THIN_BORDER
            ws.cell(row=_ROW_KEYS, column=ci, value=col).font, ws.cell(row=_ROW_KEYS, column=ci).fill, ws.cell(row=_ROW_KEYS, column=ci).alignment = _KEY_ROW_FONT, _KEY_ROW_FILL, _ALIGN_CENTER
        ws.row_dimensions[_ROW_KEYS].height = 10
        data_row = _ROW_DATA
    else:
        ws = wb[sheet_name]
        col_map = _extend_sheet_with_new_columns(ws, _read_key_row(ws), all_columns)
        data_row = ws.max_row + 1

    for col_name in all_columns:
        ci = col_map.get(col_name)
        if not ci: continue
        cell = ws.cell(row=data_row, column=ci, value=_round_value(row.get(col_name, "")))
        cell.alignment = _ALIGN_CENTER
        if col_name.startswith("delta_"):
            imp = _is_improvement(col_name, row.get(col_name), baseline_lookup.get(col_name))
            if imp is True: cell.fill, cell.font = _GREEN_FILL, _GREEN_FONT
            elif imp is False: cell.fill, cell.font = _RED_FILL, _RED_FONT
            else: cell.fill, cell.font = _NEUTRAL_FILL, _NEUTRAL_FONT

    _apply_alt_row_shading(ws, data_row)
    _autofit_columns(ws)
    _freeze_panes(ws, row=_ROW_DATA, col=4)


_EPOCH_COMPONENT_ORDER = [
    "run_id", "timestamp_utc", "epoch", "total", "ce", "dpo", "dpo_conditioned", "kl", "hinge", "unlikelihood", "entropy", "cf",
    "eval_BBQ_accuracy", "eval_StereoSet_LMS", "eval_StereoSet_ICAT", "eval_StereoSet_SS",
    "eval_CrowS_pct_stereo", "eval_CrowS_pct_anti", "eval_BOLD_regard_mean", "eval_BOLD_regard_disparity", "eval_BOLD_pct_negative",
    "eval_ToxiGen_accuracy", "eval_ToxiGen_macro_f1", "eval_ToxiGen_toxic_f1", "eval_ToxiGen_fnr_toxic", "eval_ToxiGen_accuracy_disparity",
    "eval_ToxiGen_fnr_disparity", "eval_ToxiGen_parse_fail_rate", "eval_ImplicitBBQ_accuracy", "eval_ImplicitBBQ_hedge_rate",
    "eval_FairMT_bias_rate", "eval_FairMT_bias_disparity", "eval_FairMT_parse_fail_rate",
]

def _write_epoch_curves_sheet(wb: Workbook, epoch_logs: List[Dict], run_id: str, timestamp_utc: str) -> None:
    if not epoch_logs: return
    sheet_name = "Epoch Curves"
    
    seen = {"run_id", "timestamp_utc", "epoch"}
    all_keys = list(seen) + [k for log in epoch_logs for k in sorted(log.keys()) if not (seen.add(k) or True)]
    all_keys = [k for k in _EPOCH_COMPONENT_ORDER if k in all_keys] + sorted(k for k in all_keys if k not in _EPOCH_COMPONENT_ORDER)
    is_new = sheet_name not in wb.sheetnames

    if is_new:
        ws = wb.create_sheet(sheet_name)
        for ci, col in enumerate(all_keys, 1):
            ws.cell(row=1, column=ci, value=_make_readable_metric_name(col)).font, ws.cell(row=1, column=ci).fill, ws.cell(row=1, column=ci).alignment, ws.cell(row=1, column=ci).border = _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER, _THIN_BORDER
            ws.cell(row=2, column=ci, value=col).font, ws.cell(row=2, column=ci).fill, ws.cell(row=2, column=ci).alignment = _KEY_ROW_FONT, _KEY_ROW_FILL, _ALIGN_CENTER
        ws.row_dimensions[2].height = 10
        data_start = 3
    else:
        ws = wb[sheet_name]
        existing = [ws.cell(row=2, column=ci).value for ci in range(1, ws.max_column + 1)]
        for k in all_keys:
            if k not in existing:
                ci = len(existing) + 1
                ws.cell(row=1, column=ci, value=_make_readable_metric_name(k)).font, ws.cell(row=1, column=ci).fill, ws.cell(row=1, column=ci).alignment = _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER
                ws.cell(row=2, column=ci, value=k).font, ws.cell(row=2, column=ci).fill, ws.cell(row=2, column=ci).alignment = _KEY_ROW_FONT, _KEY_ROW_FILL, _ALIGN_CENTER
                existing.append(k)
        all_keys = [k for k in existing if k]
        data_start = ws.max_row + 1

    for ri, log in enumerate(epoch_logs):
        for ci, col in enumerate(all_keys, 1):
            val = run_id if col == "run_id" else (timestamp_utc if col == "timestamp_utc" else _round_value(log.get(col, "")))
            ws.cell(row=data_start + ri, column=ci, value=val).alignment = _ALIGN_CENTER

    _apply_alt_row_shading(ws, data_start)
    _autofit_columns(ws)
    _freeze_panes(ws, row=data_start, col=3)


def _build_summary(cfg_dict: Dict) -> str:
    loss_fn = cfg_dict.get("loss_fn", "")
    parts = [f"{k}={cfg_dict[k]}" for k in ["loss_fn", "lr", "bias_lambda", "hinge_lambda", "dpo_beta", "dpo_ref", "kl_lambda", "kl_mode", "epochs", "batch_size", "use_instruction", "use_cot", "per_epoch_eval", "stereo_from_metadata", "train_dataset", "eval_datasets", "toxigen_threshold"] if cfg_dict.get(k) not in ("", None)]
    
    parts.append(f"train_instr=[{_format_prompt_list(cfg_dict.get('train_instructions', ''))}]")
    parts.append(f"eval_prompts=[{_format_prompt_list(cfg_dict.get('eval_prompts', ''))}]")
    
    if cfg_dict.get("use_iterative"):
        parts.append(f"iterative=True(reason={cfg_dict.get('iterative_max_reasoning_tokens', '?')},ans={cfg_dict.get('iterative_max_answer_tokens', '?')})")
    else:
        parts.append("iterative=False")
        
    if cfg_dict.get("fairmtbench_judge_model"):
        parts.append(f"fairmt_judge={cfg_dict['fairmtbench_judge_model']}")
        parts.append(f"fairmt_cats=[{_format_prompt_list(cfg_dict.get('fairmtbench_categories', ''))}]")

    if loss_fn in _ENTROPY_LOSSES and cfg_dict.get("entropy_lambda"): parts.append(f"entropy_λ={cfg_dict['entropy_lambda']}")
    if loss_fn in _CONDITIONED_LOSSES:
        w_parts = [f"{lbl}={cfg_dict[f]}" for f, lbl in [("cond_w_ambig_neg", "A×N"), ("cond_w_ambig_nonneg", "A×NN"), ("cond_w_disambig_neg", "D×N"), ("cond_w_disambig_nonneg", "D×NN")] if cfg_dict.get(f)]
        if w_parts: parts.append(f"cond_weights=[{', '.join(w_parts)}]")
    if loss_fn in _CF_LOSSES:
        if cfg_dict.get("cf_lambda"): parts.append(f"cf_λ={cfg_dict['cf_lambda']}")
        if cfg_dict.get("cf_margin"): parts.append(f"cf_γ={cfg_dict['cf_margin']}")
        
    return "  |  ".join(parts)


def _write_config_index_sheet(wb: Workbook, run_id: str, timestamp_utc: str, cfg_dict: Dict) -> None:
    sheet_name = "Config Index"
    all_columns = ["run_id", "timestamp_utc", "summary"] + sorted(cfg_dict.keys())
    is_new = sheet_name not in wb.sheetnames

    if is_new:
        ws = wb.create_sheet(sheet_name)
        for ci, col in enumerate(all_columns, 1):
            ws.cell(row=1, column=ci, value=_READABLE_METRIC_NAMES.get(col, col)).font, ws.cell(row=1, column=ci).fill, ws.cell(row=1, column=ci).alignment, ws.cell(row=1, column=ci).border = _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER, _THIN_BORDER
            ws.cell(row=2, column=ci, value=col).font, ws.cell(row=2, column=ci).fill, ws.cell(row=2, column=ci).alignment = _KEY_ROW_FONT, _KEY_ROW_FILL, _ALIGN_CENTER
        ws.row_dimensions[2].height = 10
        data_start = 3
    else:
        ws = wb[sheet_name]
        existing = [ws.cell(row=2, column=ci).value for ci in range(1, ws.max_column + 1)]
        for k in all_columns:
            if k not in existing:
                ci = len(existing) + 1
                ws.cell(row=1, column=ci, value=_READABLE_METRIC_NAMES.get(k, k)).font, ws.cell(row=1, column=ci).fill, ws.cell(row=1, column=ci).alignment = _HEADER_FONT, _HEADER_FILL, _ALIGN_CENTER
                ws.cell(row=2, column=ci, value=k).font, ws.cell(row=2, column=ci).fill, ws.cell(row=2, column=ci).alignment = _KEY_ROW_FONT, _KEY_ROW_FILL, _ALIGN_CENTER
                existing.append(k)
        all_columns = [k for k in existing if k]
        data_start = ws.max_row + 1

    row_data = {"run_id": run_id, "timestamp_utc": timestamp_utc, "summary": _build_summary(cfg_dict), **cfg_dict}
    for ci, col in enumerate(all_columns, 1):
        ws.cell(row=data_start, column=ci, value=row_data.get(col, "")).alignment = Alignment(wrap_text=False, vertical="center", horizontal="left")

    _apply_alt_row_shading(ws, data_start)
    _autofit_columns(ws)
    _freeze_panes(ws, row=data_start, col=3)


def _build_vertical_config_header(run_id: str, cfg_dict: Dict, result_row: Dict, timestamp_utc: str) -> Tuple[str, str, str]:
    loss_fn = cfg_dict.get("loss_fn", result_row.get("loss_fn", "?"))
    l1 = [f"Run: {run_id}", f"Model: {cfg_dict.get('model_id', '?')}", f"Loss: {loss_fn}", f"Train: {cfg_dict.get('train_dataset', '?')}", f"DPO ref: {cfg_dict.get('dpo_ref', '?')}", f"KL mode: {cfg_dict.get('kl_mode', '?')}", f"Stereo meta: {cfg_dict.get('stereo_from_metadata', '?')}", f"Instruction: {cfg_dict.get('use_instruction', '?')}", f"CoT: {cfg_dict.get('use_cot', '?')}", f"Epochs: {cfg_dict.get('epochs', '?')}", f"LR: {cfg_dict.get('lr', '?')}", f"λ_bias: {cfg_dict.get('bias_lambda', '?')}", f"λ_hinge: {cfg_dict.get('hinge_lambda', '?')}", f"β_dpo: {cfg_dict.get('dpo_beta', '?')}", f"λ_kl: {cfg_dict.get('kl_lambda', '?')}"]
    
    if loss_fn in _ENTROPY_LOSSES: l1.append(f"λ_entropy: {cfg_dict.get('entropy_lambda', '?')}")
    if loss_fn in _CONDITIONED_LOSSES: l1.append(f"cond_w: A×N={cfg_dict.get('cond_w_ambig_neg', '?')} A×NN={cfg_dict.get('cond_w_ambig_nonneg', '?')} D×N={cfg_dict.get('cond_w_disambig_neg', '?')} D×NN={cfg_dict.get('cond_w_disambig_nonneg', '?')}")
    if loss_fn in _CF_LOSSES: l1.extend([f"λ_cf: {cfg_dict.get('cf_lambda', '?')}", f"γ_cf: {cfg_dict.get('cf_margin', '?')}"])
    
    l2 = [f"Train Instructions: [{_format_prompt_list(cfg_dict.get('train_instructions', ''))}]", f"Eval Prompts: [{_format_prompt_list(cfg_dict.get('eval_prompts', ''))}]"]
    l2.append(f"Iterative Prompting: {'ON' if cfg_dict.get('use_iterative') else 'OFF'}")
    if cfg_dict.get("fairmtbench_judge_model"): l2.append(f"FairMT Judge: {cfg_dict['fairmtbench_judge_model']}")
    
    l3 = [f"Timestamp: {timestamp_utc}", f"Eval datasets: {cfg_dict.get('eval_datasets', '?')}", f"N train: {result_row.get('n_train_samples', '?')}", f"N biased: {result_row.get('n_biased_samples', '?')}", f"Wall time: {result_row.get('wall_time_s', '?')}s"]
    if cfg_dict.get("fairmtbench_categories"): l3.append(f"FairMT cats: [{_format_prompt_list(cfg_dict['fairmtbench_categories'])}]")
    
    return "   |   ".join(l1), "   |   ".join(l2), "   |   ".join(l3)


def _write_vertical_results_sheet(wb: Workbook, result_row: Dict, run_id: str, timestamp_utc: str, cfg_dict: Optional[Dict] = None) -> None:
    sheet_name = "Vertical Results"
    cfg_dict = cfg_dict or {}
    triplets = _build_triplet_columns(result_row)
    ordered_sections, _ = _build_ordered_sections(triplets, result_row)
    baseline_lookup = {t["delta_key"]: float(result_row.get(t["baseline_key"], 0)) for t in triplets if t["delta_key"] and t["baseline_key"]}
    is_new = sheet_name not in wb.sheetnames

    ws = wb.create_sheet(sheet_name) if is_new else wb[sheet_name]
    start_row = 1 if is_new else ws.max_row + 3

    line1, line2, line3 = _build_vertical_config_header(run_id, cfg_dict, result_row, timestamp_utc)
    for r, text, font, size in [(start_row, line1, _VERT_RUN_HEADER_FONT, 10), (start_row + 1, line2, Font(bold=True, color="2E75B6", size=9), 9), (start_row + 2, line3, Font(italic=True, color="444444", size=9), 9)]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=_VERT_NUM_COLS)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font, cell.fill, cell.alignment = font, _VERT_RUN_HEADER_FILL, Alignment(wrap_text=True, vertical="center", horizontal="left")

    header_row = start_row + 4
    for ci, hdr in enumerate(["Section", "Metric", "Baseline", "Post-FT", "Δ", "Direction"], 1):
        ws.cell(row=header_row, column=ci, value=hdr).font, ws.cell(row=header_row, column=ci).fill, ws.cell(row=header_row, column=ci).alignment, ws.cell(row=header_row, column=ci).border = _VERT_HEADER_FONT, _VERT_HEADER_FILL, _ALIGN_CENTER, _THIN_BORDER

    current_row = header_row + 1
    for sec in ordered_sections:
        if not sec["triplets"]: continue
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=_VERT_NUM_COLS)
        ws.cell(row=current_row, column=1, value=sec["name"]).font, ws.cell(row=current_row, column=1).fill, ws.cell(row=current_row, column=1).alignment = _VERT_SECTION_FONT, _VERT_SECTION_FILL, _ALIGN_LEFT
        current_row += 1

        for t in sec["triplets"]:
            is_primary = t["raw"] in (_HIGHER_IS_BETTER | _LOWER_IS_BETTER | _IDEAL_AT_50)
            baseline_val = _round_value(result_row.get(t["baseline_key"], "") if t["baseline_key"] else "")
            postft_val = _round_value(result_row.get(t["postft_key"], "") if t["postft_key"] else "")
            delta_val = _round_value(result_row.get(t["delta_key"], "") if t["delta_key"] else "")
            
            ws.cell(row=current_row, column=1, value="")
            ws.cell(row=current_row, column=2, value=_make_readable_metric_name(t["raw"])).font, ws.cell(row=current_row, column=2).alignment = (_VERT_METRIC_FONT_BOLD if is_primary else _VERT_METRIC_FONT), _ALIGN_LEFT
            ws.cell(row=current_row, column=3, value=baseline_val).alignment = _ALIGN_CENTER
            ws.cell(row=current_row, column=4, value=postft_val).alignment = _ALIGN_CENTER
            
            d_cell = ws.cell(row=current_row, column=5, value=delta_val)
            d_cell.alignment = _ALIGN_CENTER
            if t["delta_key"] and delta_val != "":
                imp = _is_improvement(t["delta_key"], delta_val, baseline_lookup.get(t["delta_key"]))
                if imp is True: d_cell.fill, d_cell.font = _GREEN_FILL, _GREEN_FONT
                elif imp is False: d_cell.fill, d_cell.font = _RED_FILL, _RED_FONT
                else: d_cell.fill, d_cell.font = _NEUTRAL_FILL, _NEUTRAL_FONT
                
            ws.cell(row=current_row, column=6, value=_direction_label(t["delta_key"] or t["baseline_key"] or "")).font, ws.cell(row=current_row, column=6).alignment = _VERT_DIRECTION_FONT, _ALIGN_CENTER
            current_row += 1

    _apply_alt_row_shading(ws, header_row + 1)
    for ci, w in [(1, 8), (2, 38), (3, 14), (4, 14), (5, 14), (6, 16)]:
        ws.column_dimensions[get_column_letter(ci)].width = w
    _freeze_panes(ws, row=header_row + 1, col=3)


# =============================================================================
# 8. MAIN ENTRY POINT
# =============================================================================

def append_results_xlsx(
    xlsx_path: str | Path,
    result_row: Dict,
    epoch_logs: Optional[List[Dict]] = None,
    run_id: str = "",
    timestamp_utc: str = "",
    cfg_dict: Optional[Dict] = None,
) -> None:
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    epoch_logs = epoch_logs or []
    cfg_dict = cfg_dict or {}

    wb = load_workbook(str(xlsx_path)) if xlsx_path.exists() else Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _write_results_sheet(wb, result_row)
    _write_epoch_curves_sheet(wb, epoch_logs, run_id, timestamp_utc)
    _write_config_index_sheet(wb, run_id, timestamp_utc, cfg_dict)
    _write_vertical_results_sheet(wb, result_row, run_id, timestamp_utc, cfg_dict)

    wb.save(str(xlsx_path))
    print(f"  XLSX results saved -> {xlsx_path}")